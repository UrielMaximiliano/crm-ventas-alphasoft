"""Export de leads a XLSX para que el equipo lo trabaje en Excel.

El xlsx tiene una sola hoja "Leads" con:
- Header row bold sobre fondo gris, fuente Arial 11.
- Freeze panes en A2 (al scrollear, el header queda fijo).
- Auto filter para filtrar/ordenar dentro de Excel.
- Hyperlinks en columnas "Maps URL" y "Sitio Web".
- Wrap text en las columnas largas (mensajes generados).
- Anchos de columna razonables (no auto fit dinamico, pero buenos defaults).

No usa formulas - el xlsx es solo lectura de datos, las decisiones las toma el
equipo manualmente fuera del archivo.
"""
from __future__ import annotations

import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.db.models import Lead, Message, MessageChannel


_HEADER_FILL = PatternFill("solid", start_color="1F2937")  # slate-800
_HEADER_FONT = Font(name="Arial", size=11, bold=True, color="FFFFFF")
_BODY_FONT = Font(name="Arial", size=10)
_WRAP = Alignment(wrap_text=True, vertical="top")
_TOP_ALIGN = Alignment(vertical="top")


# (header, width, wrap?)
_COLUMNS: list[tuple[str, int, bool]] = [
    ("ID", 6, False),
    ("Score", 7, False),
    ("Score Reason", 32, True),
    ("Nombre", 30, False),
    ("Rubro", 22, False),
    ("Ciudad", 18, False),
    ("Provincia", 18, False),
    ("Telefono", 18, False),
    ("Email", 30, False),
    ("Emails Extra", 30, True),
    ("Telefonos Extra", 22, True),
    ("Sitio Web", 32, False),
    ("Estado Sitio", 14, False),
    ("Direccion", 38, True),
    ("Rating", 8, False),
    ("Estado", 12, False),
    ("Razon Calificacion", 28, True),
    ("Analisis del Sitio", 50, True),
    ("Dolores Detectados", 40, True),
    ("Servicio Sugerido", 22, False),
    ("Search Query", 28, False),
    ("Mensaje WhatsApp", 60, True),
    ("Asunto Email", 36, True),
    ("Mensaje Email", 80, True),
    ("Maps URL", 30, False),
    ("Creado", 18, False),
]


def _latest_message_by_channel(lead: Lead) -> dict[str, Message]:
    out: dict[str, Message] = {}
    for m in sorted(lead.messages or [], key=lambda x: x.generated_at):
        out[m.channel.value] = m
    return out


def _fmt_dt(dt: datetime | None) -> str:
    return dt.strftime("%Y-%m-%d %H:%M") if dt else ""


def build_leads_xlsx(leads: list[Lead]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"

    # Header
    for col_idx, (name, width, _wrap) in enumerate(_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(vertical="center", horizontal="left", wrap_text=False)
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 22

    # Filas
    for r_offset, lead in enumerate(leads, start=2):
        msgs = _latest_message_by_channel(lead)
        wsp = msgs.get(MessageChannel.WHATSAPP.value)
        eml = msgs.get(MessageChannel.EMAIL.value)
        row_values: list[str | int | float | None] = [
            lead.id,
            lead.priority_score,
            lead.priority_reason,
            lead.name,
            lead.category,
            lead.city,
            lead.province,
            lead.phone,
            lead.email,
            lead.extracted_emails,
            lead.extracted_phones,
            lead.website,
            lead.website_status,
            lead.address,
            lead.rating,
            lead.status.value if lead.status else None,
            lead.qualification_reason,
            lead.site_analysis,
            lead.pain_points,
            lead.recommended_service,
            lead.search_query,
            wsp.body if wsp else None,
            eml.subject if eml else None,
            eml.body if eml else None,
            lead.source_id,
            _fmt_dt(lead.created_at),
        ]
        for c_offset, value in enumerate(row_values, start=1):
            cell = ws.cell(row=r_offset, column=c_offset, value=value)
            cell.font = _BODY_FONT
            _, _, wrap = _COLUMNS[c_offset - 1]
            cell.alignment = _WRAP if wrap else _TOP_ALIGN

        # Hyperlinks - hago el href clickable pero sigo mostrando el texto.
        # Las columnas se mueven si _COLUMNS cambia: las resolvemos por nombre.
        web_col = next(i for i, c in enumerate(_COLUMNS, start=1) if c[0] == "Sitio Web")
        maps_col = next(i for i, c in enumerate(_COLUMNS, start=1) if c[0] == "Maps URL")

        web_cell = ws.cell(row=r_offset, column=web_col)
        if lead.website and lead.website.startswith(("http://", "https://")):
            web_cell.hyperlink = lead.website
            web_cell.font = Font(name="Arial", size=10, color="1F6FEB", underline="single")

        maps_cell = ws.cell(row=r_offset, column=maps_col)
        if lead.source_id and lead.source_id.startswith(("http://", "https://")):
            maps_cell.hyperlink = lead.source_id
            maps_cell.font = Font(name="Arial", size=10, color="1F6FEB", underline="single")
            maps_cell.value = "Ver en Maps"

    # Freeze + autofilter
    ws.freeze_panes = "A2"
    last_col_letter = get_column_letter(len(_COLUMNS))
    ws.auto_filter.ref = f"A1:{last_col_letter}{max(1, len(leads) + 1)}"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_leads_filename(prefix: str = "leads") -> str:
    ts = datetime.now().strftime("%Y%m%d_%H%M")
    return f"{prefix}_alphasoft_{ts}.xlsx"
